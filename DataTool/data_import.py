# coding:utf-8
import sys
import time
from datetime import datetime
import re
import os
import json
import openpyxl
import uuid
from sqlalchemy import Column, String, DateTime, create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base


# 创建对象的基类:
Base = declarative_base()

# 定义User对象:


class User(Base):
    # 表的名字:
    __tablename__ = 'user'

    # 表的结构:
    userId = Column(String(36), primary_key=True)
    username = Column(String(20))
    email = Column(String(20))
    password = Column(String(20))
    note = Column(String(20))
    create_time = Column(DateTime(), default=datetime.now)
    updated_time = Column(
        DateTime(), default=datetime.now, onupdate=datetime.now)


class orm(object):

    def __init__(self):
        self._filepath = 'D:\\data.xlsx'
        self.wb = openpyxl.load_workbook(self._filepath)

    def redisinfo(self):
        ws = self.wb['redis']
        self.redislist = []
        for i in range(3, ws.max_column):
            self.relist = []
            self.relist.append(ws.cell(row=1, column=i).value)
            self.relist.append(ws.cell(row=2, column=i).value)
            self.relist.append(ws.cell(row=3, column=i).value)
            self.relist.append(ws.cell(row=4, column=i).value)
            self.relist.append(ws.cell(row=5, column=i).value)
            self.relist.append(ws.cell(row=6, column=i).value)
            self.relist.append(ws.cell(row=7, column=i).value)
            self.relist.append(ws.cell(row=8, column=i).value)
            self.relist.append(ws.cell(row=9, column=i).value)
            self.relist.append(ws.cell(row=10, column=i).value)
            self.relist.append(ws.cell(row=11, column=i).value)
            self.redislist.append(self.relist)
        return(self.redislist)


def save():

    name = "test_name"
    namespace = "test_namespace"

    # 初始化数据库连接:
    engine = create_engine('mysql+mysqlconnector://root:123456@localhost:3306/od_appconf', pool_recycle=3600)
    # 创建DBSession类型:
    DBSession = sessionmaker(bind=engine)

    # 创建session对象:
    session = DBSession()
    # 创建新User对象:
    new_user = User()
    new_user.userId = uuid.uuid3(namespace, name)
    new_user.username = "admin"
    new_user.email = "henu5972@163.com"
    new_user.password = "123456"
    new_user.note = "Happy day"
    # 添加到session:
    session.add(new_user)
    # 提交即保存到数据库:
    session.commit()
    # 关闭session:
    session.close()

    # 创建Session:
    session = DBSession()
    # 创建Query查询，filter是where条件，最后调用one()返回唯一行，如果调用all()则返回所有行:
    user = session.query(User).filter(User.id == '5').one()
    # 打印类型和对象的name属性:
    print('type:', type(user))
    print('name:', user.name)
    # 关闭Session:
    session.close()


if __name__ == '__main__':
    #maininfo = orm()
    #datas = maininfo.redisinfo()
    save()
