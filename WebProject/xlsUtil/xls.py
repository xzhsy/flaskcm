# coding:utf-8
from WebProject import app
import sys
import time
import re
import os
import json
import openpyxl

ipP = r"?P<ip>[\d.]*"
ndP = r"?P<node>[\d]"
grP = r"?P<group>[A-Z]"
prP = r"?P<prn>[a-zA-Z]"
zh_pattern = re.compile(u'[\u4e00-\u9fa5]+')
h = re.compile(r"(%s)" % (ipP), re.VERBOSE)
n = re.compile(r"(%s)" % (ndP), re.VERBOSE)
g = re.compile(r"(%s)" % (grP), re.VERBOSE)
p = re.compile(r"(%s)" % (prP), re.VERBOSE)


class orm(object):

    def __init__(self):
        path=os.path.join(app.root_path,'xlsData')
        self.file='test.xlsx'
        self._filepath = os.path.join(path, self.file)
        self.wb = openpyxl.load_workbook(self._filepath )
        # print(wb.sheetnames)
        self.wd = self.wb['主机汇总']
        self.project = []
        self.app = []
        self.ip = []
        self.group = []
        self.node = []
        self.applist=[]
        self.zk=[]



    def gen_port(self, app, node):
        # print(node)
        if app.find('core') > 0:
            num = 30100
        elif app.find('control') > 0:
            num = 20100
        else:
            num = 10100
        self.http = []
        self.https = []
        self.ajp = []
        self.shutdown = []
        self.jmx = []
        self.dubbo= []
        if node == None or app.find('mq') != -1:
            self.http = ['null']
            self.https = ['null']
            self.ajp = ['null']
            self.shutdown = ['null']
            self.jmx = ['null']
            self.dubbo = ['null']
        else:
            for i in range(1, int(node) + 1):
                self.http.append(num + 10 + i)
                self.https.append(num + 20 + i)
                self.ajp.append(num + 30 + i)
                self.shutdown.append(num + 40 + i)
                self.jmx.append(num + 50 + i)
                self.dubbo.append(num + 60 + i)
                print(app,num)
        return (self.http,self.https,self.ajp,self.shutdown,self.jmx,self.dubbo)

    def zkinfo(self):
        ws = self.wb['ZK规划']
        for i in range(2,ws.max_row):
            self.zklist = []
            self.zklist.append(ws.cell(row=i, column=3).value)
            self.zklist.append(ws.cell(row=i, column=4).value)
            self.zklist.append(ws.cell(row=i, column=5).value)
            self.zklist.append(ws.cell(row=i, column=8).value)
            self.zklist.append(ws.cell(row=i, column=9).value)
            self.zklist.append(ws.cell(row=i, column=10).value)
            self.zk.append(self.zklist)
        return(self.zk)
    def redisinfo(self):
        ws = self.wb['redis']
        self.redislist=[]
        for i in range(3,ws.max_column):
            self.relist = []
            self.relist.append(ws.cell(row=1, column=i).value)
            self.relist.append(ws.cell(row=2, column=i).value)
            self.relist.append(ws.cell(row=3, column=i).value)
            self.relist.append(ws.cell(row=4, column=i).value)
            self.relist.append(ws.cell(row=5, column=i).value)
            self.relist.append(ws.cell(row=6, column=i).value)
            self.relist.append(ws.cell(row=7, column=i).value)
            self.relist.append(ws.cell(row=8, column=i).value)
            self.relist.append(ws.cell(row=9, column=i).value)
            self.relist.append(ws.cell(row=10, column=i).value)
            self.relist.append(ws.cell(row=11, column=i).value)
            self.redislist.append(self.relist)
        return(self.redislist)

    def main(self, prn):
        mat = p.match(prn)
        # print(prn,mat)

        for i in range(2, self.wd.max_row):
            self.h = []
            appname = self.wd.cell(row=i, column=5).value
            gr = self.wd.cell(row=i, column=8).value
            node=self.wd.cell(row=i, column=7).value
            if node == None:
                node = 1
            app = self.wd.cell(row=i, column=6).value
            #print(app)
            if appname == None or app == None:
                continue
            if mat != None:
                if app.find(prn) == -1 and prn.find(app) == -1:
                    continue
            else:
                if appname.find(prn) == -1 and prn.find(appname) == -1:
                    continue
            print(app, node)
            if gr != 'X':

                self.gen_port(app, node)
            else:
                node = 1
                self.gen_port(app, node)
            # print(appname)
            self.h.append(appname)
            self.h.append(app)
            self.h.append(gr)
            self.h.append(self.wd.cell(row=i, column=10).value)
            self.h.append(node)
            self.h.append(self.http)
            self.h.append(self.https)
            self.h.append(self.ajp)
            self.h.append(self.shutdown)
            self.h.append(self.jmx)
            self.h.append(self.dubbo)
            self.applist.append(self.h)

        return(self.applist)

        # print(len(self.project),len(self.h))
        # return  (self.h)

if __name__ == '__main__':
    maininfo = orm(prn='ngcs')
    maininfo.redisinfo()