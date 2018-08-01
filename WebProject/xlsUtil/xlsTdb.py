# coding:utf-8
from WebProject.models.models import procfg
from WebProject.route import repository
from WebProject import app
import sys
import time
import re
import os
import json
import openpyxl
import uuid


db = repository.GetSession()

ipP = r"?P<ip>[\d.]*"
ndP = r"?P<node>[\d]"
grP = r"?P<group>[A-Z]"
prP = r"?P<prn>[a-zA-Z]"
zh_pattern = re.compile(u'[\u4e00-\u9fa5]+')
h = re.compile(r"(%s)" % (ipP), re.VERBOSE)
n = re.compile(r"(%s)" % (ndP), re.VERBOSE)
g = re.compile(r"(%s)" % (grP), re.VERBOSE)
p = re.compile(r"(%s)" % (prP), re.VERBOSE)


class xlsTdb(object):

    def __init__(self):
        path = os.path.join(app.root_path, 'xlsData')
        self.file = 'test.xlsx'
        self._filepath = os.path.join(path, self.file)
        self.wb = openpyxl.load_workbook(self._filepath)
        # print(wb.sheetnames)
        self.wd = self.wb['主机汇总']
        self.project = ''
        self.app = ''
        self.ip = ' '
        self.group = ' '
        self.node = ''
        self.applist = ' '
        self.zk = []

    def gen_port(self, app, node):
        # print(node)
        if app.find('core') > 0:
            num = 30100
        elif app.find('control') > 0:
            num = 20100
        else:
            num = 10100
        self.http = ' '
        self.https = ' '
        self.ajp = ' '
        self.shutdown = ' '
        self.jmx = ' '
        self.dubbo = ' '
        if node == None or 'mq' in app or app == None:
            self.http = 'null'
            self.https = 'null'
            self.ajp = 'null'
            self.shutdown = 'null'
            self.jmx = 'null'
            self.dubbo = 'null'
        else:
            # for i in range(1, int(node) + 1):
            self.http = num + 10 + node
            self.https = num + 20 + node
            self.ajp = num + 30 + node
            self.shutdown = num + 40 + node
            self.jmx = num + 50 + node
            self.dubbo = num + 60 + node
        return (self.http,self.https,self.ajp,self.shutdown,self.jmx,self.dubbo)

    def zkinfo(self):
        ws = self.wb['ZK规划']
        for i in range(2,ws.max_row):
            self.zklist = []
            self.zklist.append(ws.cell(row=i, column=3).value)
            self.zklist.append(ws.cell(row=i, column=4).value)
            self.zklist.append(ws.cell(row=i, column=5).value)
            self.zklist.append(ws.cell(row=i, column=8).value)
            self.zklist.append(ws.cell(row=i, column=9).value)
            self.zklist.append(ws.cell(row=i, column=10).value)
            self.zk.append(self.zklist)
        return(self.zk)
    def redisinfo(self):
        ws = self.wb['redis']
        self.redislist=[]
        for i in range(3,ws.max_column):
            self.relist = []
            self.relist.append(ws.cell(row=1, column=i).value)
            self.relist.append(ws.cell(row=2, column=i).value)
            self.relist.append(ws.cell(row=3, column=i).value)
            self.relist.append(ws.cell(row=4, column=i).value)
            self.relist.append(ws.cell(row=5, column=i).value)
            self.relist.append(ws.cell(row=6, column=i).value)
            self.relist.append(ws.cell(row=7, column=i).value)
            self.relist.append(ws.cell(row=8, column=i).value)
            self.relist.append(ws.cell(row=9, column=i).value)
            self.relist.append(ws.cell(row=10, column=i).value)
            self.relist.append(ws.cell(row=11, column=i).value)
            self.redislist.append(self.relist)
        return(self.redislist)

    def prj2data(self, filename):
        self.file=filename
        # print(prn,mat)

        for i in range(2, self.wd.max_row):
            self.h = ' '
            appname = self.wd.cell(row=i, column=5).value
            gr = self.wd.cell(row=i, column=8).value
            node=self.wd.cell(row=i, column=7).value
            if node == None:
                node = 1
            app = self.wd.cell(row=i, column=6).value
            # print(app)
            if not app:
                app= ''
            # print(app, node)
            ip = self.wd.cell(row=i, column=10).value            
            # print(proid)    
            if gr == 'X':
                node = 1
            # print(app, node)    
            for i in range(1, int(node) + 1):
                proid = str(uuid.uuid4())
                xls=procfg()
                xls.gr=gr
                xls.node=node
                xls.proname=appname
                xls.project=app
                xls.proid=proid
                self.gen_port(app, i)
                xls.http=self.http
                xls.https=self.https
                xls.ajp=self.ajp
                xls.jmx=self.jmx
                xls.IPAddress=ip
                xls.dubbo=self.dubbo
                xls.shutdown=self.shutdown
                # print(xls.http)
                db.add(xls)
        db.commit()
        db.close() 

    def main(self,filename):
        asd=xlsTdb()
        asd.prj2data(filename)

    if __name__ == '__main__':
        main('test.xlsx')           
